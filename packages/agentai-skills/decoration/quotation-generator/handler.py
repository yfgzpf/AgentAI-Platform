#!/usr/bin/env python3
"""
Quotation Generator - 智能装修报价生成器
"""

import json
import sys
import os
from datetime import datetime


# 价格数据库
PRICE_DB = {
    "labor": {
        "一线城市": {"拆除": 80, "水电": 150, "瓦工": 120, "木工": 100, "油漆": 60},
        "二线城市": {"拆除": 60, "水电": 120, "瓦工": 100, "木工": 80, "油漆": 50},
        "三线城市": {"拆除": 40, "水电": 90, "瓦工": 80, "木工": 60, "油漆": 40},
    },
    "material": {
        "经济": {"瓷砖": 60, "地板": 80, "涂料": 20, "吊顶": 50, "橱柜": 800, "卫浴": 3000},
        "舒适": {"瓷砖": 120, "地板": 200, "涂料": 40, "吊顶": 100, "橱柜": 1500, "卫浴": 6000},
        "豪华": {"瓷砖": 300, "地板": 500, "涂料": 100, "吊顶": 250, "橱柜": 3500, "卫浴": 15000},
    }
}


def parse_input(input_data: dict) -> dict:
    """解析输入参数"""
    prompt = input_data.get("prompt", "")
    params = input_data.get("params", {})
    
    # 解析面积
    area_match = None
    import re
    area_pattern = r'(\d+)\s*平[方米]?'
    match = re.search(area_pattern, prompt)
    if match:
        area_match = int(match.group(1))
    
    # 解析户型
    room_pattern = r'(\d+)\s*室\s*(\d+)\s*厅\s*(\d+)\s*卫'
    room_match = re.search(room_pattern, prompt)
    rooms = room_match.group(0) if room_match else "3室2厅1卫"
    
    # 解析风格
    styles = ["现代简约", "北欧", "中式", "轻奢", "美式", "日式"]
    style = "现代简约"
    for s in styles:
        if s in prompt:
            style = s
            break
    
    # 解析档次
    if "豪华" in prompt or "高端" in prompt:
        level = "豪华"
    elif "经济" in prompt or "便宜" in prompt or "省钱" in prompt:
        level = "经济"
    else:
        level = "舒适"
    
    # 城市级别
    tier1 = ["北京", "上海", "广州", "深圳"]
    tier2 = ["杭州", "南京", "武汉", "成都", "西安", "重庆", "天津", "苏州"]
    
    city = params.get("city", "")
    if any(c in city for c in tier1):
        city_level = "一线城市"
    elif any(c in city for c in tier2):
        city_level = "二线城市"
    else:
        city_level = "三线城市"
    
    return {
        "area": area_match or params.get("area", 100),
        "rooms": rooms,
        "style": style,
        "level": level,
        "city_level": city_level,
        "special_items": params.get("special_items", [])
    }


def calculate_quotation(params: dict) -> dict:
    """计算报价"""
    area = params["area"]
    level = params["level"]
    city_level = params["city_level"]
    
    # 获取价格
    labor_prices = PRICE_DB["labor"][city_level]
    material_prices = PRICE_DB["material"][level]
    
    # 计算各项目
    breakdown = {}
    
    # 基础工程
    base_labor = (labor_prices["拆除"] + labor_prices["水电"] + 
                  labor_prices["瓦工"] + labor_prices["油漆"]) * area
    base_material = (material_prices["涂料"] * 3 + 50) * area  # 墙面+地面
    breakdown["基础工程"] = {
        "labor": base_labor,
        "material": base_material,
        "total": base_labor + base_material
    }
    
    # 客餐厅（按30%面积）
    living_area = area * 0.3
    living_labor = (labor_prices["瓦工"] + labor_prices["木工"] + labor_prices["油漆"]) * living_area
    living_material = (material_prices["瓷砖"] + material_prices["吊顶"] + 200) * living_area
    breakdown["客餐厅"] = {
        "labor": living_labor,
        "material": living_material,
        "total": living_labor + living_material
    }
    
    # 卧室（按40%面积，3个房间）
    bedroom_area = area * 0.4
    bedroom_labor = (labor_prices["地板"] + labor_prices["油漆"]) * bedroom_area
    bedroom_material = (material_prices["地板"] + 100) * bedroom_area  # 地板+衣柜
    breakdown["卧室"] = {
        "labor": bedroom_labor,
        "material": bedroom_material,
        "total": bedroom_labor + bedroom_material
    }
    
    # 厨房（按10%面积）
    kitchen_area = area * 0.1
    kitchen_labor = labor_prices["瓦工"] * kitchen_area + 3000  # 橱柜安装
    kitchen_material = material_prices["瓷砖"] * kitchen_area + material_prices["橱柜"]
    breakdown["厨房"] = {
        "labor": kitchen_labor,
        "material": kitchen_material,
        "total": kitchen_labor + kitchen_material
    }
    
    # 卫生间（按12%面积，2个）
    bath_area = area * 0.12
    bath_labor = labor_prices["瓦工"] * bath_area + 2000
    bath_material = material_prices["瓷砖"] * bath_area + material_prices["卫浴"]
    breakdown["卫生间"] = {
        "labor": bath_labor,
        "material": bath_material,
        "total": bath_labor + bath_material
    }
    
    # 其他（门、窗、开关、五金、管理费等）
    other_labor = 2000
    other_material = area * 80 + 5000  # 门+窗+开关+五金
    breakdown["其他"] = {
        "labor": other_labor,
        "material": other_material,
        "total": other_labor + other_material
    }
    
    # 特殊项目
    special_total = 0
    for item in params.get("special_items", []):
        if item == "地暖":
            special_total += area * 150
        elif item == "新风":
            special_total += 15000
        elif item == "中央空调":
            special_total += area * 300
        elif item == "智能家居":
            special_total += 20000
    
    if special_total > 0:
        breakdown["特殊项目"] = {
            "labor": 0,
            "material": special_total,
            "total": special_total
        }
    
    # 计算总计
    total_labor = sum(v["labor"] for v in breakdown.values())
    total_material = sum(v["material"] for v in breakdown.values())
    total = sum(v["total"] for v in breakdown.values())
    
    # 管理费（10%）
    management_fee = int(total * 0.1)
    total_with_management = total + management_fee
    
    return {
        "total": total_with_management,
        "per_sqm": int(total_with_management / area),
        "area": area,
        "labor_total": total_labor,
        "material_total": total_material,
        "management_fee": management_fee,
        "breakdown": breakdown,
        "params": params
    }


def generate_excel(quotation: dict, customer_info: dict = None) -> str:
    """生成Excel报价单"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "装修报价单"
        
        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = "装修工程报价单"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        ws['A3'] = "客户信息"
        ws['A3'].font = Font(bold=True)
        if customer_info:
            ws['A4'] = f"姓名: {customer_info.get('name', '')}"
            ws['A5'] = f"电话: {customer_info.get('phone', '')}"
        ws['A6'] = f"面积: {quotation['area']}㎡"
        ws['A7'] = f"档次: {quotation['params']['level']}"
        ws['A8'] = f"日期: {datetime.now().strftime('%Y-%m-%d')}"
        
        # 表头
        headers = ["项目", "人工费", "材料费", "小计", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据
        row = 11
        for project, values in quotation['breakdown'].items():
            ws.cell(row=row, column=1, value=project)
            ws.cell(row=row, column=2, value=values['labor'])
            ws.cell(row=row, column=3, value=values['material'])
            ws.cell(row=row, column=4, value=values['total'])
            row += 1
        
        # 管理费
        ws.cell(row=row, column=1, value="管理费(10%)")
        ws.cell(row=row, column=4, value=quotation['management_fee'])
        row += 1
        
        # 总计
        ws.cell(row=row, column=1, value="总计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=quotation['total'])
        ws.cell(row=row, column=4).font = Font(bold=True, size=14)
        
        # 保存
        os.makedirs("output", exist_ok=True)
        filename = f"output/quotation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
        
    except ImportError:
        return None
    except Exception as e:
        print(f"Excel生成失败: {e}", file=sys.stderr)
        return None


def main():
    try:
        input_data = json.load(sys.stdin)
        
        # 解析参数
        params = parse_input(input_data)
        
        # 计算报价
        quotation = calculate_quotation(params)
        
        # 生成Excel
        customer_info = input_data.get("customer", {})
        excel_file = generate_excel(quotation, customer_info)
        
        # 构建输出
        output_text = f"""📊 装修报价单

🏠 项目信息:
  • 面积: {params['area']}㎡
  • 户型: {params['rooms']}
  • 风格: {params['style']}
  • 档次: {params['level']}
  • 城市: {params['city_level']}

💰 费用明细:
"""
        
        for project, values in quotation['breakdown'].items():
            output_text += f"  • {project}: ¥{values['total']:,} (人工¥{values['labor']:,} + 材料¥{values['material']:,})\n"
        
        output_text += f"""
📋 费用汇总:
  • 人工费合计: ¥{quotation['labor_total']:,}
  • 材料费合计: ¥{quotation['material_total']:,}
  • 管理费(10%): ¥{quotation['management_fee']:,}
  ─────────────────
  • 总报价: ¥{quotation['total']:,}
  • 单价: ¥{quotation['per_sqm']}/㎡

💡 说明:
  1. 以上价格为估算，实际以现场测量为准
  2. 不含家具、家电、窗帘等软装
  3. 付款方式: 签约30% → 水电完成30% → 泥木完成30% → 验收10%
"""
        
        if excel_file:
            output_text += f"\n📄 Excel报价单已生成: {excel_file}"
        
        result = {
            "success": True,
            "output": output_text,
            "data": {
                **quotation,
                "excel_file": excel_file
            }
        }
        
        print(json.dumps(result))
        
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)


if __name__ == "__main__":
    main()
