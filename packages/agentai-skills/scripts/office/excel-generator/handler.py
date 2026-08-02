#!/usr/bin/env python3
"""
Excel Generator Handler
从 JSON 数据自动生成 Excel 表格 (基于 openpyxl)
"""

import json
import sys
import os
import io

def generate_excel(data, config):
    """生成 Excel 并返回 bytes"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        return None, "请安装 openpyxl: pip install openpyxl"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.get('sheet_name', 'Sheet1')

    # 写入数据
    if isinstance(data, list):
        if len(data) == 0:
            return None, "数据为空"

        # 表头
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row, item in enumerate(data, 2):
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(h, ''))
        elif isinstance(data[0], list):
            for row, item in enumerate(data, 1):
                for col, val in enumerate(item, 1):
                    ws.cell(row=row, column=col, value=val)

    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # 添加图表 (如果数据包含数值列)
    chart_config = config.get('chart', {})
    if chart_config.get('enabled'):
        chart = BarChart()
        chart.type = 'col'
        chart.title = chart_config.get('title', '数据图表')
        chart.y_axis.title = chart_config.get('y_label', '')
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data)+1, max_col=2)
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, 'E5')

    # 写入 buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), None


def main():
    try:
        input_data = json.load(sys.stdin)
        data = input_data.get('data', [])
        config = input_data.get('config', {})
        filename = input_data.get('filename', 'output.xlsx')

        if not data:
            print(json.dumps({
                'success': False,
                'output': 'Missing required parameter: data (list of dicts or list of lists)'
            }))
            return

        file_bytes, error = generate_excel(data, config)

        if error:
            print(json.dumps({
                'success': False,
                'output': error
            }))
            return

        # 保存文件
        output_path = config.get('output_path', os.path.join(os.getcwd(), filename))
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

        with open(output_path, 'wb') as f:
            f.write(file_bytes)

        size_kb = len(file_bytes) / 1024
        print(json.dumps({
            'success': True,
            'output': f"Excel 生成成功: {output_path} ({size_kb:.1f} KB, {len(data)} 行)",
            'data': {
                'path': output_path,
                'rows': len(data),
                'size_kb': round(size_kb, 1)
            }
        }))

    except Exception as e:
        print(json.dumps({
            'success': False,
            'output': f'Error: {str(e)}'
        }))

if __name__ == '__main__':
    main()
